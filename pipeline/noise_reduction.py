"""
2.2 数据清洗与去重 (Noise Reduction - Content-Based)

根据 PRD 需求：
- 数据展平与文本拼接
- TF-IDF + 余弦相似度计算（字符级 n-gram）
- 94% 阈值去重（保留策略：最新文件 > 最多字符）
- 时间线容错重构
"""
from __future__ import annotations

import logging
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import numpy as np
from openpyxl.worksheet.worksheet import Worksheet
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

from pipeline.models import (
    DuplicateGroup,
    SheetRecord,
)
from pipeline.utils import parse_date_from_text

logger = logging.getLogger(__name__)


# ============================================================================
# Step 1: 数据展平与文本拼接
# ============================================================================

def flatten_sheet_to_text(ws: Worksheet) -> str:
    """
    将一个 Sheet 的核心业务字段展平为纯文本字符串。

    识别"本周完成工作内容"和"下周工作计划"两个区域，
    将每行的 任务描述 + 进度 + 难点分析 拼接为文本行。

    Args:
        ws: openpyxl Worksheet 对象

    Returns:
        拼接后的纯文本字符串
    """
    text_parts = []
    current_section = None

    for row_idx in range(1, (ws.max_row or 0) + 1):
        # 读取 A 列判断区域
        col_a = ws.cell(row_idx, 1).value
        if col_a is not None:
            col_a_str = str(col_a).strip()
            if "本周完成" in col_a_str or "工作内容" in col_a_str:
                current_section = "tasks"
                continue
            elif "下周" in col_a_str and ("计划" in col_a_str or "工作" in col_a_str):
                current_section = "plans"
                continue
            elif col_a_str in ("日期", "序号"):
                continue

        if current_section is None:
            continue

        # 收集 C 列（任务描述/内容）和 E 列（难点/描述）
        col_c = ws.cell(row_idx, 3).value
        col_d = ws.cell(row_idx, 4).value
        col_e = ws.cell(row_idx, 5).value

        if col_c is None:
            continue

        parts = [str(col_c).strip()]

        # 进度（仅任务区域有意义）
        if current_section == "tasks" and col_d is not None:
            try:
                progress = float(col_d)
                parts.append(f"进度:{progress}")
            except (ValueError, TypeError):
                parts.append(str(col_d).strip())

        # 难点分析/描述
        if col_e is not None:
            e_str = str(col_e).strip()
            if e_str:
                parts.append(e_str)

        line = " | ".join(parts)
        if line.strip():
            text_parts.append(line)

    return "\n".join(text_parts)


# ============================================================================
# Step 2: TF-IDF + 余弦相似度计算
# ============================================================================

def compute_similarity_matrix(texts: list[str]) -> np.ndarray:
    """
    使用 TF-IDF + 余弦相似度计算文本间相似度矩阵。

    采用字符级 n-gram (2-4 gram) 策略，适合中文技术术语。
    无需额外分词依赖，且对嵌入式领域专业术语更鲁棒。

    Args:
        texts: 纯文本字符串列表

    Returns:
        N×N 的相似度矩阵 (numpy ndarray)
    """
    n = len(texts)

    if n == 0:
        return np.array([]).reshape(0, 0)

    if n == 1:
        return np.array([[1.0]])

    # 处理空字符串：替换为占位符以避免 TfidfVectorizer 报错
    processed_texts = [t if t.strip() else " " for t in texts]

    try:
        vectorizer = TfidfVectorizer(
            analyzer='char_wb',
            ngram_range=(2, 4),    # 2-4 字符的 n-gram
            max_features=10000,    # 限制特征数，加速大文本计算
            sublinear_tf=True,     # 使用对数 TF，减少高频词影响
        )
        tfidf_matrix = vectorizer.fit_transform(processed_texts)
        sim_matrix = cosine_similarity(tfidf_matrix)
    except ValueError as e:
        # 所有文本都为空或过短时的兜底
        logger.warning("TF-IDF 向量化失败: %s，返回单位矩阵", e)
        sim_matrix = np.eye(n)

    return sim_matrix


# ============================================================================
# Step 3: 94% 阈值去重
# ============================================================================

def _build_duplicate_groups(
    records: list[SheetRecord],
    sim_matrix: np.ndarray,
    threshold: float,
) -> list[set[int]]:
    """
    基于相似度矩阵和阈值，使用 Union-Find 构建重复组。

    将相似度 ≥ threshold 的 Sheet 索引归入同一连通分量。

    Returns:
        重复组列表，每组为索引集合
    """
    n = len(records)
    parent = list(range(n))

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]  # 路径压缩
            x = parent[x]
        return x

    def union(x: int, y: int):
        px, py = find(x), find(y)
        if px != py:
            parent[px] = py

    # 合并相似度 ≥ threshold 的对
    for i in range(n):
        for j in range(i + 1, n):
            if sim_matrix[i][j] >= threshold:
                union(i, j)

    # 收集连通分量
    groups: dict[int, set[int]] = defaultdict(set)
    for i in range(n):
        groups[find(i)].add(i)

    # 只返回大小 > 1 的组（即有重复的组）
    return [group for group in groups.values() if len(group) > 1]


def _select_survivor(records: list[SheetRecord], indices: set[int]) -> int:
    """
    在重复组中选择"优胜者"。

    优先规则（PRD 保留策略）：
    1. file_modified_time 最新的
    2. char_count 最多的

    Returns:
        优胜者在 records 中的索引
    """
    candidates = list(indices)

    # 按 (修改时间降序, 字符数降序) 排序，取第一个
    def sort_key(idx: int):
        r = records[idx]
        mtime = r.file_modified_time or datetime.min
        return (-mtime.timestamp(), -r.char_count)

    candidates.sort(key=sort_key)
    return candidates[0]


def deduplicate_sheets(
    records: list[SheetRecord],
    similarity_threshold: float = 0.98,
) -> tuple[list[SheetRecord], list[DuplicateGroup]]:
    """
    对同一员工的 Sheet 列表进行基于内容的去重。

    Args:
        records: 同一员工的所有 SheetRecord
        similarity_threshold: 相似度阈值，默认 0.98

    Returns:
        (survivors, duplicate_groups)
        - survivors: 去重后保留的 SheetRecord 列表
        - duplicate_groups: 被判定为重复的组信息
    """
    if len(records) <= 1:
        return list(records), []

    # 计算相似度矩阵
    texts = [r.raw_text for r in records]
    sim_matrix = compute_similarity_matrix(texts)

    # 构建重复组
    dup_groups_indices = _build_duplicate_groups(records, sim_matrix, similarity_threshold)

    # 确定保留和丢弃
    discard_indices: set[int] = set()
    dup_groups: list[DuplicateGroup] = []

    for group_indices in dup_groups_indices:
        survivor_idx = _select_survivor(records, group_indices)
        discarded_indices = group_indices - {survivor_idx}
        discard_indices.update(discarded_indices)

        # 收集组内相似度分数
        group_list = sorted(group_indices)
        scores = []
        for i, idx_i in enumerate(group_list):
            for idx_j in group_list[i + 1:]:
                scores.append(float(sim_matrix[idx_i][idx_j]))

        dup_groups.append(DuplicateGroup(
            survivor=records[survivor_idx],
            discarded=[records[i] for i in sorted(discarded_indices)],
            similarity_scores=scores,
        ))

    # 构建去重后的列表
    survivors = [r for i, r in enumerate(records) if i not in discard_indices]

    logger.info(
        "去重完成: %d 份输入 → %d 份保留, %d 个重复组",
        len(records), len(survivors), len(dup_groups),
    )

    return survivors, dup_groups


# ============================================================================
# Step 4: 时间线容错重构
# ============================================================================

def reconstruct_timeline(records: list[SheetRecord]) -> list[SheetRecord]:
    """
    对去重后的 Sheet 按时间正序重新排列。

    优先使用已解析的 date_range，若无则降级使用 file_modified_time。

    Args:
        records: 去重后的 SheetRecord 列表

    Returns:
        按时间正序排列的 SheetRecord 列表
    """
    def sort_key(r: SheetRecord):
        # 优先使用 date_range 的起始日期
        if r.date_range is not None:
            return r.date_range[0]
        # 降级使用文件修改时间（转为 date）
        if r.file_modified_time is not None:
            return r.file_modified_time.date()
        # 兜底使用一个极早的日期
        return date.min

    return sorted(records, key=sort_key)


# ============================================================================
# 高层编排函数
# ============================================================================

def flatten_and_deduplicate(
    all_records: list[SheetRecord],
    similarity_threshold: float = 0.98,
) -> tuple[dict[str, list[SheetRecord]], list[DuplicateGroup]]:
    """
    对全量 SheetRecord 按员工分组 → 去重 → 时间线重构。

    Args:
        all_records: 来自 auto_discovery 的全部 SheetRecord
        similarity_threshold: 相似度阈值

    Returns:
        (employee_timelines, all_duplicate_groups)
        - employee_timelines: {员工邮箱: [按时间排序的 SheetRecord]}
        - all_duplicate_groups: 所有重复组信息
    """
    # 按员工分组
    by_employee: dict[str, list[SheetRecord]] = defaultdict(list)
    for r in all_records:
        by_employee[r.employee_email].append(r)

    employee_timelines: dict[str, list[SheetRecord]] = {}
    all_dup_groups: list[DuplicateGroup] = []

    for emp_email, emp_records in sorted(by_employee.items()):
        logger.info("处理员工 %s: %d 份 Sheet", emp_email, len(emp_records))

        # 去重
        survivors, dup_groups = deduplicate_sheets(emp_records, similarity_threshold)
        all_dup_groups.extend(dup_groups)

        # 时间线重构
        timeline = reconstruct_timeline(survivors)
        employee_timelines[emp_email] = timeline

        logger.info(
            "  → %d 份保留, %d 个重复组剔除",
            len(timeline), len(dup_groups),
        )

    return employee_timelines, all_dup_groups
