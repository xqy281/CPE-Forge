"""
公共工具函数 — 文件解析、日期提取、员工名映射等

提供管线各模块共用的底层能力，包含完善的异常处理和容错机制。
"""
from __future__ import annotations

import logging
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)

# ============================================================================
# TSD（腾讯文档）格式标识
# ============================================================================
TSD_MAGIC_BYTES = b'%TSD'

# ============================================================================
# 文件加载
# ============================================================================

def detect_file_format(filepath: Path) -> str:
    """
    通过文件头魔数检测文件实际格式。

    Returns:
        "xlsx"  — ZIP-based Office Open XML
        "xls"   — OLE2 Compound Document (旧版 Excel)
        "tsd"   — 腾讯文档 (Tencent Spreadsheet Document)
        "unknown" — 无法识别的格式
    """
    try:
        with open(filepath, 'rb') as fh:
            header = fh.read(8)
    except (OSError, IOError) as e:
        logger.error("无法读取文件头: %s, 错误: %s", filepath, e)
        return "unknown"

    if header[:4] == b'\x50\x4b\x03\x04':
        return "xlsx"
    elif header[:4] == b'\xd0\xcf\x11\xe0':
        return "xls"
    elif header[:4] == TSD_MAGIC_BYTES:
        return "tsd"
    else:
        return "unknown"


def safe_load_workbook(filepath: Path) -> tuple[Optional[Workbook], str, str]:
    """
    容错加载 Excel 工作簿，优先 openpyxl，兜底 xlrd。

    对于 TSD 格式或无法识别的格式，返回 None 并标注原因。

    Returns:
        (workbook_or_none, file_format, error_message)
        - workbook: 成功时返回 openpyxl Workbook，失败返回 None
        - file_format: 检测到的文件格式 ("xlsx"/"xls"/"tsd"/"unknown")
        - error_message: 错误信息，成功时为空字符串
    """
    fmt = detect_file_format(filepath)

    # TSD 格式 — 腾讯文档加密/私有格式
    if fmt == "tsd":
        return None, fmt, "TSD 格式（腾讯文档），疑似加密文件"

    # 标准 xlsx 格式
    if fmt == "xlsx":
        try:
            wb = openpyxl.load_workbook(str(filepath), data_only=True)
            return wb, fmt, ""
        except Exception as e:
            return None, fmt, f"openpyxl 加载失败: {e}"

    # 旧版 xls 格式 — 使用 xlrd 尝试读取
    if fmt == "xls":
        try:
            import xlrd
            xls_wb = xlrd.open_workbook(str(filepath))
            # 将 xlrd workbook 转换为 openpyxl 兼容的结构
            # 我们创建一个内存中的 openpyxl workbook 来统一接口
            wb = _convert_xlrd_to_openpyxl(xls_wb)
            xls_wb.release_resources()
            return wb, fmt, ""
        except Exception as e:
            err_str = str(e).lower()
            if "encrypted" in err_str or "password" in err_str:
                return None, fmt, f"加密的 xls 文件: {e}"
            return None, fmt, f"xlrd 加载失败: {e}"

    # 未知格式 — 尝试所有方式
    # 先尝试 openpyxl
    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True)
        return wb, "xlsx", ""
    except Exception:
        pass

    # 再尝试 xlrd
    try:
        import xlrd
        xls_wb = xlrd.open_workbook(str(filepath))
        wb = _convert_xlrd_to_openpyxl(xls_wb)
        xls_wb.release_resources()
        return wb, "xls", ""
    except Exception:
        pass

    # 检查是否为乱码（疑似加密）
    try:
        with open(filepath, 'rb') as fh:
            sample = fh.read(512)
        # 如果前 512 字节中可打印的 ASCII 字符占比很低，视为加密
        printable_ratio = sum(1 for b in sample if 32 <= b < 127) / max(len(sample), 1)
        if printable_ratio < 0.3:
            return None, "unknown", "文件内容疑似加密（可打印字符占比过低）"
    except Exception:
        pass

    return None, "unknown", "无法识别的文件格式，所有解析器均失败"


def _convert_xlrd_to_openpyxl(xls_wb) -> Workbook:
    """
    将 xlrd Workbook 转换为 openpyxl Workbook，统一后续处理接口。

    仅转换文本值，不处理公式和样式。
    """
    wb = openpyxl.Workbook()
    # 删除默认创建的 sheet
    wb.remove(wb.active)

    for sheet_idx in range(xls_wb.nsheets):
        xls_ws = xls_wb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=xls_ws.name)

        for row_idx in range(xls_ws.nrows):
            for col_idx in range(xls_ws.ncols):
                cell_value = xls_ws.cell_value(row_idx, col_idx)
                if cell_value != '':
                    ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

    return wb


# ============================================================================
# 日期提取
# ============================================================================

# 匹配模式：2025年1月6日、01月12日、1/6、01.06 等
_DATE_PATTERNS = [
    # 完整格式：2025年1月6日 或 2025年01月06日（「日」可选）
    re.compile(
        r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?'
    ),
    # 短格式（无年份）：1月6日 或 01月06日（「日」可选）
    re.compile(
        r'(\d{1,2})\s*月\s*(\d{1,2})\s*日?'
    ),
    # 斜杠格式：1/6、01/06
    re.compile(
        r'(\d{1,2})\s*/\s*(\d{1,2})'
    ),
    # 点分格式：01.06
    re.compile(
        r'(\d{2})\.(\d{2})'
    ),
]

# 匹配日期区间：起始日期 ~ 结束日期（分隔符支持 ~、-、—、——、至）
_RANGE_SEPARATORS = re.compile(r'[~\-—]+|至')


def parse_date_from_text(text: str, reference_year: int = 2025) -> Optional[tuple[date, date]]:
    """
    从文本中提取日期区间。

    按照 PRD 要求，优先从表头/内容中正则匹配日期。
    支持多种日期格式和分隔符。

    Args:
        text: 待提取的文本
        reference_year: 当年份缺失时的默认参考年份

    Returns:
        (start_date, end_date) 或 None
    """
    if not text or not text.strip():
        return None

    # 智能推断 reference_year：如果文本中出现了完整年份，则优先使用
    # 例如 "工作周报_何宗峰(2024年12月30日-2025年1月3日)" 中可提取到 2024
    year_match = re.search(r'(\d{4})\s*年', text)
    if year_match:
        reference_year = int(year_match.group(1))

    # 尝试找到日期区间（含分隔符的）
    # 先尝试匹配 "2025年1月6日~1月11日" 这种完整区间
    range_match = re.search(
        r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?'
        r'\s*[~\-—]+\s*'
        r'(?:(\d{4})\s*年\s*)?(\d{1,2})\s*月\s*(\d{1,2})\s*日?',
        text
    )
    if range_match:
        g = range_match.groups()
        y1 = int(g[0])
        m1, d1 = int(g[1]), int(g[2])
        y2 = int(g[3]) if g[3] else y1
        m2, d2 = int(g[4]), int(g[5])
        # 处理跨年（12月~1月）
        if m2 < m1 and y2 == y1:
            y2 += 1
        try:
            return (date(y1, m1, d1), date(y2, m2, d2))
        except ValueError:
            pass

    # 尝试匹配 "1月6日~1月11日" 或 "1/6 ~ 1/11"（无年份）
    range_match_short = re.search(
        r'(\d{1,2})\s*[月/]\s*(\d{1,2})\s*日?\s*'
        r'[~\-—]+\s*'
        r'(\d{1,2})\s*[月/]\s*(\d{1,2})\s*日?',
        text
    )
    if range_match_short:
        g = range_match_short.groups()
        m1, d1 = int(g[0]), int(g[1])
        m2, d2 = int(g[2]), int(g[3])
        y = reference_year
        if m2 < m1:
            try:
                return (date(y, m1, d1), date(y + 1, m2, d2))
            except ValueError:
                pass
        else:
            try:
                return (date(y, m1, d1), date(y, m2, d2))
            except ValueError:
                pass

    # 匹配 "01.06-01.11" 格式
    dot_range = re.search(
        r'(\d{2})\.(\d{2})\s*-\s*(\d{2})\.(\d{2})',
        text
    )
    if dot_range:
        g = dot_range.groups()
        m1, d1 = int(g[0]), int(g[1])
        m2, d2 = int(g[2]), int(g[3])
        y = reference_year
        try:
            return (date(y, m1, d1), date(y, m2, d2))
        except ValueError:
            pass

    # 只匹配到单个日期时，返回以该日期为起止的区间
    for pattern in _DATE_PATTERNS:
        match = pattern.search(text)
        if match:
            groups = match.groups()
            try:
                if len(groups) == 3:
                    # 完整日期含年份
                    d = date(int(groups[0]), int(groups[1]), int(groups[2]))
                    return (d, d)
                elif len(groups) == 2:
                    d = date(reference_year, int(groups[0]), int(groups[1]))
                    return (d, d)
            except ValueError:
                continue

    return None


def extract_employee_name_from_filename(filename: str) -> str:
    """
    从文件名中提取员工中文姓名。

    支持的格式：
    - "工作周报_萧倩云(2025年...)" → "萧倩云"
    - "萧倩云软件部2025年..." → "萧倩云"
    - "吴开健_软件部（2025年...）..." → "吴开健"
    - "工作周报_赖灿辉（2025年...）" → "赖灿辉"
    """
    # 模式1：工作周报_姓名(日期) 或 工作周报_姓名（日期）
    m = re.match(r'工作周报[_\s]*([^\(（_\d]+)', filename)
    if m:
        return m.group(1).strip()

    # 模式2：姓名_软件部 或 姓名软件部
    m = re.match(r'([^\d_]+?)[_]?软件部', filename)
    if m:
        return m.group(1).strip()

    # 兜底：取文件名前面的中文字符
    m = re.match(r'([\u4e00-\u9fff]+)', filename)
    if m:
        return m.group(1)

    return ""


def get_file_modified_time(filepath: Path) -> Optional[datetime]:
    """获取文件的最后修改时间"""
    try:
        mtime = os.path.getmtime(str(filepath))
        return datetime.fromtimestamp(mtime)
    except OSError:
        return None
