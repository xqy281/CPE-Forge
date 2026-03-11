"""
2.1 智能识别与过滤 (Auto-Discovery)

根据 PRD 需求：
- 不依赖文件名，读取 Excel 表头
- 若包含预设特征列命中率超过阈值，则判定为有效周报
- 剔除无关财务/行政表格
- TSD 格式文件标记为加密文件（而非丢弃）
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from pipeline.eml_extractor import calibrate_year

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from pipeline.models import (
    FileResult,
    FileStatus,
    SheetRecord,
    TaskRow,
    PlanRow,
)
from pipeline.utils import (
    detect_file_format,
    extract_employee_name_from_filename,
    get_file_modified_time,
    parse_date_from_text,
    safe_load_workbook,
)

logger = logging.getLogger(__name__)

# ============================================================================
# 预设特征列 — 用于识别周报 Sheet
# ============================================================================
FEATURE_COLUMNS = {
    "任务描述", "进度", "难点", "下周计划", "工作内容",
    "本周完成", "序号", "日期", "难点分析", "心得",
    "工作总结", "计划时间", "内容", "描述", "总结",
    "下周工作", "完成工作",
}

# 部分特征可以是子串匹配（如"难点分析/详细描述/总结/心得"）
FEATURE_SUBSTRINGS = [
    "任务描述", "进度", "难点", "下周", "工作内容",
    "本周完成", "序号", "日期", "心得", "总结",
    "计划时间", "工作计划",
]

# 默认匹配阈值
DEFAULT_THRESHOLD = 0.6


def is_weekly_report_sheet(
    ws: Worksheet,
    feature_columns: set[str] | None = None,
    threshold: float = DEFAULT_THRESHOLD,
) -> tuple[bool, float]:
    """
    判断一个 Sheet 是否为有效周报。

    通过读取前 3 行的所有单元格值，与预设特征列进行匹配。
    采用子串匹配策略，以应对"难点分析/详细描述/总结/心得"这类复合表头。

    Args:
        ws: openpyxl Worksheet 对象
        feature_columns: 自定义特征列集合（默认使用 FEATURE_SUBSTRINGS）
        threshold: 匹配阈值，默认 0.6

    Returns:
        (is_valid, match_score)
        - is_valid: 是否判定为有效周报
        - match_score: 匹配率 (0.0 ~ 1.0)
    """
    if feature_columns is None:
        substrings = FEATURE_SUBSTRINGS
    else:
        substrings = list(feature_columns)

    # 收集前 3 行的所有单元格文本
    cell_texts = set()
    try:
        for row_idx in range(1, min(ws.max_row or 0, 3) + 1):
            for col_idx in range(1, min(ws.max_column or 0, 20) + 1):
                val = ws.cell(row_idx, col_idx).value
                if val is not None:
                    cell_texts.add(str(val).strip())
    except Exception as e:
        logger.warning("读取 Sheet '%s' 表头失败: %s", ws.title, e)
        return False, 0.0

    if not cell_texts:
        return False, 0.0

    # 合并所有单元格文本为一个大字符串，用于子串匹配
    combined_text = " ".join(cell_texts)

    # 统计命中的特征子串数
    matched = sum(1 for sub in substrings if sub in combined_text)
    total = len(substrings)

    if total == 0:
        return False, 0.0

    score = matched / total
    return score >= threshold, score


def _parse_sheet_records(
    ws: Worksheet,
    employee_name: str,
    employee_email: str,
    source_file: Path,
    file_modified_time: str,
    calibration_map: dict[str, datetime] | None = None,
) -> Optional[SheetRecord]:
    """
    解析单个 Sheet 为 SheetRecord。

    识别"本周完成工作内容"和"下周工作计划"两个区域，逐行提取任务和计划。

    Returns:
        SheetRecord 或 None（如果 Sheet 数据不足）
    """
    tasks = []
    plans = []
    current_section = None  # "tasks" 或 "plans"
    date_str_from_content = ""

    try:
        for row_idx in range(1, (ws.max_row or 0) + 1):
            # 读取 A 列判断区域切换
            col_a = ws.cell(row_idx, 1).value
            if col_a is not None:
                col_a_str = str(col_a).strip()
                if "本周完成" in col_a_str or "工作内容" in col_a_str:
                    current_section = "tasks"
                    continue
                elif "下周" in col_a_str and ("计划" in col_a_str or "工作" in col_a_str):
                    current_section = "plans"
                    continue
                elif col_a_str in ("日期", "序号"):
                    continue  # 表头行，跳过
                else:
                    # 可能是日期字符串（如 "1/6 ~ 1/11"）
                    if current_section and not date_str_from_content:
                        date_str_from_content = col_a_str

            # 读取数据行
            col_b = ws.cell(row_idx, 2).value  # 序号
            col_c = ws.cell(row_idx, 3).value  # 任务描述/内容
            col_d = ws.cell(row_idx, 4).value  # 进度/计划时间
            col_e = ws.cell(row_idx, 5).value  # 难点/描述

            # 跳过空行（序号和内容都为空）
            if col_b is None and col_c is None:
                continue

            if current_section == "tasks" and col_c is not None:
                try:
                    progress = float(col_d) if col_d is not None else None
                except (ValueError, TypeError):
                    progress = None
                try:
                    seq_val = int(col_b) if col_b is not None else len(tasks) + 1
                except (ValueError, TypeError):
                    seq_val = len(tasks) + 1
                tasks.append(TaskRow(
                    seq=seq_val,
                    description=str(col_c).strip(),
                    progress=progress,
                    analysis=str(col_e).strip() if col_e else "",
                ))
            elif current_section == "plans" and col_c is not None:
                try:
                    seq_val = int(col_b) if col_b is not None else len(plans) + 1
                except (ValueError, TypeError):
                    seq_val = len(plans) + 1
                plans.append(PlanRow(
                    seq=seq_val,
                    content=str(col_c).strip(),
                    planned_time=str(col_d).strip() if col_d else "",
                    description=str(col_e).strip() if col_e else "",
                ))
    except Exception as e:
        logger.warning("解析 Sheet '%s' 数据失败: %s", ws.title, e)
        return None

    if not tasks and not plans:
        return None

    # 拼接纯文本
    text_parts = []
    for t in tasks:
        parts = [t.description]
        if t.analysis:
            parts.append(t.analysis)
        text_parts.append(" | ".join(parts))
    for p in plans:
        parts = [p.content]
        if p.description:
            parts.append(p.description)
        text_parts.append(" | ".join(parts))

    raw_text = "\n".join(text_parts)

    # 尝试提取日期区间 — 优先级：Sheet名 > 文件名 > 内容A列
    # Sheet名对多Sheet文件最精确（如 "何宗峰软件部2026年1月12日-...工作总结"）
    # 文件名适用于单Sheet文件（Sheet名通常只是「周报」）
    # A 列日期可能是「包含前两周」制度中最早的周，不代表当前 Sheet 真实周期
    date_range = None
    # ① 最高优先级：从 Sheet 名称提取
    date_range = parse_date_from_text(ws.title)
    # ② 其次从文件名提取（单Sheet文件 Sheet名为「周报」时有效）
    if date_range is None:
        date_range = parse_date_from_text(source_file.stem)
    # ③ 最后从内容中的日期字段提取
    if date_range is None and date_str_from_content:
        date_range = parse_date_from_text(date_str_from_content)

    # 年份校准：用邮件发送日期修正文件名中的错误年份
    if date_range is not None and calibration_map:
        start_date, end_date = date_range
        corrected_year = calibrate_year(
            start_date.year, source_file.name, calibration_map,
            parsed_month=start_date.month,
        )
        if corrected_year != start_date.year:
            year_delta = corrected_year - start_date.year
            start_date = start_date.replace(year=start_date.year + year_delta)
            end_date = end_date.replace(year=end_date.year + year_delta)
            date_range = (start_date, end_date)

    return SheetRecord(
        employee_name=employee_name,
        employee_email=employee_email,
        source_file=source_file,
        sheet_name=ws.title,
        date_range=date_range,
        tasks=tasks,
        plans=plans,
        raw_text=raw_text,
        char_count=len(raw_text.replace(" ", "").replace("\n", "")),
        file_modified_time=file_modified_time,
    )


def scan_directory(
    attachments_dir: Path,
    threshold: float = DEFAULT_THRESHOLD,
    specific_email: str = None,
    calibration_map: dict[str, datetime] | None = None,
) -> tuple[list[FileResult], list[FileResult], list[FileResult]]:
    """
    扫描附件目录，识别有效周报文件。

    遍历 attachments_dir 下所有员工子目录中的 .xlsx 文件，
    对每个文件的每个 Sheet 进行表头特征匹配。

    Args:
        attachments_dir: 附件根目录（包含员工子目录）
        threshold: 匹配阈值
        specific_email: 可选，限定只扫描某一个员工的文件夹名称
        calibration_map: 可选，EML 邮件日期校准映射表

    Returns:
        (valid_files, rejected_files, error_files)
        - valid_files: 至少含 1 个有效周报 Sheet 的文件
        - rejected_files: 所有 Sheet 均未通过匹配的文件
        - error_files: 无法打开/解析的文件（含加密标记）
    """
    valid_files: list[FileResult] = []
    rejected_files: list[FileResult] = []
    error_files: list[FileResult] = []

    # 遍历员工子目录
    for emp_dir in sorted(attachments_dir.iterdir()):
        if not emp_dir.is_dir():
            continue

        employee_email = emp_dir.name
        if specific_email and employee_email != specific_email:
            continue

        for filepath in sorted(emp_dir.glob("*.xlsx")):
            # 提取员工姓名
            employee_name = extract_employee_name_from_filename(filepath.name)
            file_mtime = get_file_modified_time(filepath)

            # 容错加载
            wb, fmt, error_msg = safe_load_workbook(filepath)

            if wb is None:
                # 文件无法打开
                status = FileStatus.ENCRYPTED if fmt == "tsd" else FileStatus.CORRUPT
                if "加密" in error_msg or "encrypted" in error_msg.lower():
                    status = FileStatus.ENCRYPTED

                error_files.append(FileResult(
                    filepath=filepath,
                    status=status,
                    error_message=error_msg,
                    file_format=fmt,
                ))
                logger.info(
                    "[%s] %s: %s — %s",
                    status.value.upper(), emp_dir.name, filepath.name, error_msg
                )
                continue

            # 扫描每个 Sheet
            valid_sheets: list[SheetRecord] = []
            best_score = 0.0

            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    is_valid, score = is_weekly_report_sheet(ws, threshold=threshold)
                    best_score = max(best_score, score)

                    if is_valid:
                        record = _parse_sheet_records(
                            ws, employee_name, employee_email,
                            filepath, file_mtime,
                            calibration_map=calibration_map,
                        )
                        if record is not None:
                            valid_sheets.append(record)
            finally:
                wb.close()

            if valid_sheets:
                valid_files.append(FileResult(
                    filepath=filepath,
                    status=FileStatus.VALID,
                    sheets=valid_sheets,
                    match_score=best_score,
                    file_format=fmt,
                ))
            else:
                rejected_files.append(FileResult(
                    filepath=filepath,
                    status=FileStatus.REJECTED,
                    match_score=best_score,
                    file_format=fmt,
                ))
                logger.info(
                    "[REJECTED] %s/%s (best_score=%.2f)",
                    emp_dir.name, filepath.name, best_score,
                )

    return valid_files, rejected_files, error_files
