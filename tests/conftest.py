"""
Pytest Fixtures — 测试数据工厂

创建内存中的 openpyxl Workbook 来模拟各种 Excel 周报格式，
避免测试依赖真实文件系统。
"""
from __future__ import annotations

import os
import tempfile
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def tmp_dir(tmp_path):
    """提供一个干净的临时目录"""
    return tmp_path


def create_weekly_report_workbook(
    date_str: str = "1/6 ~ 1/11",
    tasks: list[tuple] | None = None,
    plans: list[tuple] | None = None,
    sheet_name: str = "周报",
) -> openpyxl.Workbook:
    """
    创建一个标准格式的周报 Workbook。

    Args:
        date_str: 日期区间字符串
        tasks: 任务列表, 每项为 (序号, 任务描述, 进度, 难点分析)
        plans: 计划列表, 每项为 (序号, 内容, 计划时间, 描述)
        sheet_name: Sheet 名称

    Returns:
        openpyxl.Workbook
    """
    if tasks is None:
        tasks = [
            (1, "优化MP4播放，启用三重缓冲显存", 1.0, "播放720P以下视频已无卡顿掉帧现象"),
            (2, "25P01 艾为音频版效果评估", 1.0, "音效方面优于当前的景唯方案"),
            (3, "修复网口link监听服务适配问题", 0.8, "默认指向MT7531"),
        ]
    if plans is None:
        plans = [
            (1, "合并openwrt-23.5 SDK分支", "一周", "需要到月中才能释放"),
            (2, "WiFi kite7200 适配调试", "二周", "基于R22 SDK适配"),
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 本周完成工作内容 区域
    ws.cell(1, 1, "本周完成工作内容")
    ws.cell(2, 1, "日期")
    ws.cell(2, 2, "序号")
    ws.cell(2, 3, "任务描述")
    ws.cell(2, 4, "进度")
    ws.cell(2, 5, "难点分析/详细描述/总结/心得")

    ws.cell(3, 1, date_str)
    for i, (seq, desc, progress, analysis) in enumerate(tasks, start=3):
        ws.cell(i, 2, seq)
        ws.cell(i, 3, desc)
        ws.cell(i, 4, progress)
        ws.cell(i, 5, analysis)

    # 下周工作计划 区域
    plan_start_row = len(tasks) + 5
    ws.cell(plan_start_row, 1, "下周工作计划")
    ws.cell(plan_start_row + 1, 1, "日期")
    ws.cell(plan_start_row + 1, 2, "序号")
    ws.cell(plan_start_row + 1, 3, "内容")
    ws.cell(plan_start_row + 1, 4, "计划时间")
    ws.cell(plan_start_row + 1, 5, "描述")

    for i, (seq, content, time, desc) in enumerate(plans, start=plan_start_row + 2):
        ws.cell(i, 2, seq)
        ws.cell(i, 3, content)
        ws.cell(i, 4, time)
        ws.cell(i, 5, desc)

    return wb


def save_workbook_to_dir(
    wb: openpyxl.Workbook,
    directory: Path,
    filename: str,
    modified_time: datetime | None = None,
) -> Path:
    """
    将 Workbook 保存到指定目录，并可选设置修改时间。
    """
    filepath = directory / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(filepath))

    if modified_time:
        mtime = modified_time.timestamp()
        os.utime(str(filepath), (mtime, mtime))

    return filepath


def create_non_report_workbook() -> openpyxl.Workbook:
    """
    创建一个非周报格式的 Workbook（例如：功能状态表）
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.cell(1, 1, "25HD06待开发功能状态")
    ws.cell(2, 1, "类别")
    ws.cell(2, 2, "子类")
    ws.cell(2, 3, "功能项")
    ws.cell(2, 4, "开发人")
    ws.cell(2, 5, "状态")
    ws.cell(3, 1, "项目工程")
    ws.cell(3, 2, "创建工程")
    ws.cell(3, 3, "T830仓库创建该产品配置")
    ws.cell(3, 4, "萧倩云")
    ws.cell(3, 5, "已完成")

    return wb


def create_tsd_file(directory: Path, filename: str) -> Path:
    """
    创建一个模拟的 TSD（腾讯文档）文件，用于测试检测逻辑。
    """
    filepath = directory / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)
    # TSD 文件头 + 填充二进制数据
    with open(filepath, 'wb') as f:
        f.write(b'%TSD-Header-Content-')
        f.write(bytes(range(256)) * 60)  # 填充约 16KB
    return filepath


@pytest.fixture
def standard_report_wb():
    """标准格式周报 Workbook fixture"""
    return create_weekly_report_workbook()


@pytest.fixture
def multi_sheet_report_dir(tmp_path):
    """
    模拟"包含前两周记录"制度的多 Sheet 文件目录。
    创建 3 个文件（周次递增），最新的文件包含前两周的 Sheet。
    """
    emp_dir = tmp_path / "testuser@jointelli.com"
    emp_dir.mkdir()

    # 第 1 周（单独文件，仅 1 个 sheet）
    wb1 = create_weekly_report_workbook(
        date_str="1/6 ~ 1/11",
        tasks=[(1, "WiFi驱动调试mt7993", 0.8, "MTK SDK接口变更")],
        plans=[(1, "继续WiFi驱动", "一周", "")],
        sheet_name="周报",
    )
    save_workbook_to_dir(
        wb1, emp_dir,
        "工作周报_测试员(2025年1月6日~1月11日).xlsx",
        modified_time=datetime(2025, 1, 11, 18, 0),
    )

    # 第 2 周（包含第 1 周的重复 Sheet）
    wb2 = create_weekly_report_workbook(
        date_str="1/13 ~ 1/17",
        tasks=[(1, "PCIe枚举失败定位", 1.0, "根因是BAR空间配置错误")],
        plans=[(1, "提交PCIe修复patch", "3天", "")],
        sheet_name="1月13日~1月17日",
    )
    # 复制第 1 周内容到第 2 周文件中
    ws_old = wb2.create_sheet("1月6日~1月11日", 0)
    ws_old.cell(1, 1, "本周完成工作内容")
    ws_old.cell(2, 1, "日期")
    ws_old.cell(2, 2, "序号")
    ws_old.cell(2, 3, "任务描述")
    ws_old.cell(2, 4, "进度")
    ws_old.cell(2, 5, "难点分析/详细描述/总结/心得")
    ws_old.cell(3, 1, "1/6 ~ 1/11")
    ws_old.cell(3, 2, 1)
    ws_old.cell(3, 3, "WiFi驱动调试mt7993")
    ws_old.cell(3, 4, 0.8)
    ws_old.cell(3, 5, "MTK SDK接口变更")

    save_workbook_to_dir(
        wb2, emp_dir,
        "工作周报_测试员(2025年1月13日~1月17日).xlsx",
        modified_time=datetime(2025, 1, 17, 18, 0),
    )

    # 第 3 周（包含第 1、2 周的重复 Sheet）
    wb3 = create_weekly_report_workbook(
        date_str="1/20 ~ 1/24",
        tasks=[(1, "EMMC分区表优化", 0.9, "对齐4K扇区边界")],
        plans=[(1, "EMMC压测验证", "一周", "")],
        sheet_name="1月20日~1月24日",
    )
    # 复制前两周
    for name, date_s, task_desc in [
        ("1月6日~1月11日", "1/6 ~ 1/11", "WiFi驱动调试mt7993"),
        ("1月13日~1月17日", "1/13 ~ 1/17", "PCIe枚举失败定位"),
    ]:
        ws_c = wb3.create_sheet(name, 0)
        ws_c.cell(1, 1, "本周完成工作内容")
        ws_c.cell(2, 1, "日期")
        ws_c.cell(2, 2, "序号")
        ws_c.cell(2, 3, "任务描述")
        ws_c.cell(2, 4, "进度")
        ws_c.cell(2, 5, "难点分析/详细描述/总结/心得")
        ws_c.cell(3, 1, date_s)
        ws_c.cell(3, 2, 1)
        ws_c.cell(3, 3, task_desc)
        ws_c.cell(3, 4, 1.0)
        ws_c.cell(3, 5, "已解决")

    save_workbook_to_dir(
        wb3, emp_dir,
        "工作周报_测试员(2025年1月20日~1月24日).xlsx",
        modified_time=datetime(2025, 1, 24, 18, 0),
    )

    return emp_dir
