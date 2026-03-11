"""
智能识别与过滤模块 — 单元测试

覆盖 PRD 2.1 节需求：
- 有效周报 Sheet 的识别（表头特征匹配）
- 非周报文件的拒绝
- 命中率阈值边界测试
- 损坏/TSD 格式文件的容错处理
- 多 Sheet 文件中的部分匹配
"""
from __future__ import annotations

from pathlib import Path

import pytest

from tests.conftest import (
    create_non_report_workbook,
    create_tsd_file,
    create_weekly_report_workbook,
    save_workbook_to_dir,
)
from pipeline.models import FileStatus


class TestSheetRecognition:
    """表头特征匹配识别测试"""

    def test_valid_weekly_report_detected(self, tmp_path):
        """含标准表头的 Sheet 应被识别为有效周报"""
        from pipeline.auto_discovery import is_weekly_report_sheet

        wb = create_weekly_report_workbook()
        ws = wb.active
        is_valid, score = is_weekly_report_sheet(ws)
        assert is_valid is True
        assert score >= 0.6

    def test_non_report_rejected(self, tmp_path):
        """非周报文件（功能状态表）应被拒绝"""
        from pipeline.auto_discovery import is_weekly_report_sheet

        wb = create_non_report_workbook()
        ws = wb.active
        is_valid, score = is_weekly_report_sheet(ws)
        assert is_valid is False

    def test_threshold_boundary_at_60_percent(self, tmp_path):
        """恰好达到 60% 阈值时应判定为有效"""
        from pipeline.auto_discovery import is_weekly_report_sheet, FEATURE_SUBSTRINGS
        import openpyxl

        # FEATURE_SUBSTRINGS 有 12 项，需至少 8 项命中 → 66.7% ≥ 60%
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "本周完成工作内容")  # 命中: 本周完成, 工作内容
        ws.cell(2, 1, "日期")              # 命中: 日期
        ws.cell(2, 2, "序号")              # 命中: 序号
        ws.cell(2, 3, "任务描述")          # 命中: 任务描述
        ws.cell(2, 4, "进度")              # 命中: 进度
        ws.cell(2, 5, "难点分析")          # 命中: 难点
        ws.cell(3, 1, "下周工作计划")      # 命中: 下周, 工作计划, 计划时间

        is_valid, score = is_weekly_report_sheet(ws)
        assert is_valid is True
        assert score >= 0.6

    def test_threshold_boundary_below_60_percent(self, tmp_path):
        """未达到 60% 阈值时应判定为无效"""
        from pipeline.auto_discovery import is_weekly_report_sheet
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        # 仅放 2 个匹配项 → 远低于 60%
        ws.cell(1, 1, "标题")
        ws.cell(2, 1, "日期")
        ws.cell(2, 2, "负责人")
        ws.cell(2, 3, "备注说明")

        is_valid, score = is_weekly_report_sheet(ws)
        assert is_valid is False
        assert score < 0.6

    def test_empty_sheet_rejected(self, tmp_path):
        """空 Sheet 应被拒绝且不崩溃"""
        from pipeline.auto_discovery import is_weekly_report_sheet
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active  # 完全空白

        is_valid, score = is_weekly_report_sheet(ws)
        assert is_valid is False
        assert score == 0.0


class TestFileScanning:
    """文件级别扫描测试"""

    def test_scan_valid_report_file(self, tmp_path):
        """有效周报文件整体扫描"""
        from pipeline.auto_discovery import scan_directory

        emp_dir = tmp_path / "test@jointelli.com"
        emp_dir.mkdir()
        wb = create_weekly_report_workbook()
        save_workbook_to_dir(wb, emp_dir, "工作周报_测试(2025年1月6日).xlsx")

        valid, rejected, errors = scan_directory(tmp_path)
        assert len(valid) == 1
        assert valid[0].status == FileStatus.VALID

    def test_scan_rejects_non_report(self, tmp_path):
        """非周报文件应被过滤到 rejected 列表"""
        from pipeline.auto_discovery import scan_directory

        emp_dir = tmp_path / "test@jointelli.com"
        emp_dir.mkdir()
        wb = create_non_report_workbook()
        save_workbook_to_dir(wb, emp_dir, "25hd06待开发功能状态.xlsx")

        valid, rejected, errors = scan_directory(tmp_path)
        assert len(valid) == 0
        assert len(rejected) == 1
        assert rejected[0].status == FileStatus.REJECTED

    def test_scan_handles_tsd_file(self, tmp_path):
        """TSD 格式文件应标记为 ENCRYPTED 而非崩溃"""
        from pipeline.auto_discovery import scan_directory

        emp_dir = tmp_path / "test@jointelli.com"
        emp_dir.mkdir()
        create_tsd_file(emp_dir, "工作周报_测试(2025年8月25日).xlsx")

        valid, rejected, errors = scan_directory(tmp_path)
        assert len(valid) == 0
        # TSD 文件应出现在错误列表中并标记为加密
        assert len(errors) == 1
        assert errors[0].status == FileStatus.ENCRYPTED

    def test_scan_corrupt_file_no_crash(self, tmp_path):
        """完全损坏的文件不应导致扫描崩溃"""
        from pipeline.auto_discovery import scan_directory

        emp_dir = tmp_path / "test@jointelli.com"
        emp_dir.mkdir()
        corrupt_file = emp_dir / "broken.xlsx"
        corrupt_file.write_bytes(b"totally not an excel file at all")

        valid, rejected, errors = scan_directory(tmp_path)
        assert len(valid) == 0
        assert len(errors) >= 1

    def test_multi_sheet_partial_match(self, tmp_path):
        """含 Chart Sheet + 周报 Sheet 的文件，仅周报 Sheet 被识别"""
        from pipeline.auto_discovery import scan_directory

        emp_dir = tmp_path / "test@jointelli.com"
        emp_dir.mkdir()

        wb = create_weekly_report_workbook(sheet_name="周报")
        # 增加一个 Chart sheet（非周报）
        ws_chart = wb.create_sheet("Chart1", 0)
        ws_chart.cell(1, 1, "图表数据")
        ws_chart.cell(2, 1, "X轴")
        ws_chart.cell(2, 2, "Y轴")

        save_workbook_to_dir(wb, emp_dir, "工作周报_田智辉(2025年9月8日).xlsx")

        valid, rejected, errors = scan_directory(tmp_path)
        assert len(valid) == 1
        # 应只识别出 "周报" sheet，不包含 "Chart1"
        assert len(valid[0].sheets) == 1
        assert valid[0].sheets[0].sheet_name == "周报"

    def test_scan_mixed_directory(self, tmp_path):
        """混合目录中应正确分类有效/无效/错误文件"""
        from pipeline.auto_discovery import scan_directory

        emp_dir = tmp_path / "test@jointelli.com"
        emp_dir.mkdir()

        # 1 个有效周报
        wb1 = create_weekly_report_workbook()
        save_workbook_to_dir(wb1, emp_dir, "工作周报_测试(2025年1月6日).xlsx")
        # 1 个非周报
        wb2 = create_non_report_workbook()
        save_workbook_to_dir(wb2, emp_dir, "功能状态.xlsx")
        # 1 个 TSD
        create_tsd_file(emp_dir, "加密文件.xlsx")

        valid, rejected, errors = scan_directory(tmp_path)
        assert len(valid) == 1
        assert len(rejected) == 1
        assert len(errors) == 1
