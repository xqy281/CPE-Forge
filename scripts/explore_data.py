"""探索 Excel 周报数据结构的脚本"""
import openpyxl
import os
from pathlib import Path

ATTACHMENTS_DIR = Path(r"f:\Project\CPE-Forge\attachments")

def explore_file(filepath):
    """详细查看一个 Excel 文件的内容"""
    print(f"\n{'='*80}")
    print(f"FILE: {filepath.name}")
    print(f"SIZE: {filepath.stat().st_size} bytes")
    wb = openpyxl.load_workbook(str(filepath))
    print(f"SHEETS: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  --- Sheet: '{sheet_name}' ---")
        print(f"  Dimensions: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
        for r in range(1, min(ws.max_row + 1, 30)):
            row_data = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is not None:
                    val_str = repr(v)
                    if len(val_str) > 60:
                        val_str = val_str[:60] + "..."
                    row_data.append(f"C{c}={val_str}")
            if row_data:
                print(f"    R{r}: {' | '.join(row_data)}")
    wb.close()

# 探索 萧倩云 的两种格式文件
print("\n" + "="*80)
print("EXPLORING FORMAT A: 工作周报 format")
print("="*80)

dir_a = ATTACHMENTS_DIR / "xiaoqianyun@jointelli.com"
files_a = sorted(dir_a.glob("工作周报_*.xlsx"))
if files_a:
    explore_file(files_a[0])
    if len(files_a) > 1:
        explore_file(files_a[1])

print("\n\n" + "="*80)
print("EXPLORING FORMAT B: 软件部工作总结 format")
print("="*80)

files_b = sorted(dir_a.glob("萧倩云软件部*.xlsx"))
if files_b:
    explore_file(files_b[0])
    if len(files_b) > 1:
        explore_file(files_b[1])

# 看看其他员工的文件名格式
print("\n\n" + "="*80)
print("OTHER EMPLOYEES - FILE NAME PATTERNS")
print("="*80)
for emp_dir in sorted(ATTACHMENTS_DIR.iterdir()):
    if emp_dir.is_dir():
        files = sorted(emp_dir.glob("*.xlsx"))
        print(f"\n{emp_dir.name} ({len(files)} files):")
        # Show first 3 filenames
        for f in files[:3]:
            print(f"  - {f.name}")
        if len(files) > 3:
            print(f"  ... and {len(files)-3} more")

# 查看多 sheet 文件来理解去重需求
print("\n\n" + "="*80)
print("FILES WITH MULTIPLE SHEETS")
print("="*80)
for emp_dir in sorted(ATTACHMENTS_DIR.iterdir()):
    if not emp_dir.is_dir():
        continue
    for f in sorted(emp_dir.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(str(f))
            if len(wb.sheetnames) > 1:
                print(f"  {emp_dir.name}/{f.name}: sheets={wb.sheetnames}")
            wb.close()
        except Exception as e:
            print(f"  ERROR reading {f.name}: {e}")
